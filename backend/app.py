import os
import json
import base64
import anthropic
import fitz  # PyMuPDF
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
from datetime import datetime

app = Flask(__name__)
CORS(app)

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

SYSTEM_PROMPT = """You are a CRE (commercial real estate) data extraction specialist. You will receive text extracted from a comp sheet PDF — it may be from JLL, CBRE, Cushman & Wakefield, or any custom broker format.

YOUR JOB: Extract every deal/property row and return a JSON array. Each element = one deal.

CRITICAL SPLITTING RULES — apply to every format:
1. "Property Name" and "Property Address" are often in the same visual column (name on top, address below). Split into SEPARATE fields: "Property Name" and "Property Address".
2. "Market" and "Submarket" are often in the same column (submarket in parentheses below market). Split into SEPARATE "Market" and "Submarket" fields — strip all parentheses.
3. "Sale Price" and "Price PSF" are often combined (PSF in parentheses below price). Split into SEPARATE "Sale Price" and "Price PSF" fields — strip all parentheses.
4. ANY other combined columns — split them the same way.
5. Strip ALL parentheses from all values.

ADDITIONAL RULES:
- Use clear human-readable key names with spaces (e.g. "Property Name", "Cap Rate", "Year Built")
- Extract EVERY field visible: building specs, tenant info, WALT, clear height, dock doors, % leased, # of tenants, etc.
- For Comments/bullet points: concatenate all bullet points into one string separated by " | "
- Use null for fields genuinely not present
- Return ONLY a raw JSON array — no markdown, no explanation, no preamble"""

USER_PROMPT = """Extract every row of deal data from this CRE comp sheet text. Apply all splitting rules (Property Name/Address, Market/Submarket, Sale Price/PSF). Return a JSON array — one object per deal — with all fields as separate clean keys. Strip all parentheses from values. Return only raw JSON."""


def pdf_to_text(pdf_bytes):
    """Extract full text from all pages of a PDF using PyMuPDF."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages_text = []
    for i, page in enumerate(doc):
        text = page.get_text().strip()
        if text:
            pages_text.append(f"=== PAGE {i+1} ===\n{text}")
    return "\n\n".join(pages_text)


def parse_json_response(text):
    """Robustly parse JSON array from Claude's response."""
    # Strip markdown fences
    clean = text.replace("```json", "").replace("```", "").strip()
    # Find array bounds
    s, e = clean.find("["), clean.rfind("]")
    if s != -1 and e != -1:
        try:
            parsed = json.loads(clean[s:e+1])
            if isinstance(parsed, list) and len(parsed) > 0:
                return parsed
        except json.JSONDecodeError:
            pass
    # Fallback: single object
    s, e = clean.find("{"), clean.rfind("}")
    if s != -1 and e != -1:
        try:
            obj = json.loads(clean[s:e+1])
            if isinstance(obj, dict):
                return [obj]
        except json.JSONDecodeError:
            pass
    raise ValueError("Could not parse JSON from Claude response")


def normalize_row(raw):
    """Normalize keys and clean null values."""
    out = {}
    for k, v in raw.items():
        clean_key = k.strip()
        if v is None or str(v).strip().lower() in ("null", "n/a", "none", "—", "-", ""):
            out[clean_key] = ""
        else:
            out[clean_key] = str(v).strip()
    return out


def extract_comps_from_text(text):
    """Send extracted PDF text to Claude for comp extraction."""
    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4000,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": USER_PROMPT + "\n\n" + text
        }]
    )
    response_text = "".join(b.text for b in message.content if hasattr(b, "text"))
    rows = parse_json_response(response_text)
    return [normalize_row(r) for r in rows]


def build_excel(all_rows, all_columns):
    """Build a formatted Excel workbook from extracted data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comps"

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", start_color="1F3864")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=9)
    ca = Alignment(vertical="top", wrap_text=False)
    caw = Alignment(vertical="top", wrap_text=True)
    alt = PatternFill("solid", start_color="EEF2F7")
    wht = PatternFill("solid", start_color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header: # + Source File + all discovered columns
    header = ["#", "Source File"] + all_columns
    for ci, cn in enumerate(header, 1):
        c = ws.cell(row=1, column=ci, value=cn)
        c.font = hf
        c.fill = hfill
        c.alignment = ha
        c.border = bdr
    ws.row_dimensions[1].height = 28

    # Data rows
    for ri, row in enumerate(all_rows, 2):
        fill = alt if ri % 2 == 0 else wht
        ws.cell(row=ri, column=1, value=ri - 1).font = cf
        ws.cell(row=ri, column=1).fill = fill
        ws.cell(row=ri, column=1).border = bdr
        ws.cell(row=ri, column=1).alignment = ca

        ws.cell(row=ri, column=2, value=row.get("__source", "")).font = cf
        ws.cell(row=ri, column=2).fill = fill
        ws.cell(row=ri, column=2).border = bdr
        ws.cell(row=ri, column=2).alignment = ca

        for ci, col in enumerate(all_columns, 3):
            val = row.get(col, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = cf
            c.fill = fill
            c.border = bdr
            c.alignment = caw if "comment" in col.lower() else ca
        ws.row_dimensions[ri].height = 42

    # Column widths
    special_widths = {
        "Property Name": 28, "Property Address": 32, "Address": 30,
        "Market": 16, "Submarket": 20, "Sale Date": 10, "SF": 12,
        "Sale Price": 16, "Price": 16, "Price PSF": 10, "PSF": 10,
        "Cap Rate": 11, "Seller": 24, "Buyer": 24, "Year Built": 10,
        "Clear Height": 11, "WALT": 10, "# of Tenants": 10,
        "% Leased": 9, "# of Bldgs": 8, "# of Buildings": 10,
        "Comments": 55, "Comment": 55,
    }
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    for ci, col in enumerate(all_columns, 3):
        w = special_widths.get(col, max(len(col) + 2, 12))
        ws.column_dimensions[get_column_letter(ci)].width = min(w, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    return wb


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "message": "CompIQ Extractor is running"})


@app.route("/extract", methods=["POST"])
def extract():
    """
    Accepts one or more PDF files, extracts comp data, returns JSON.
    Form data: files[] = PDF file(s)
    """
    if "files[]" not in request.files:
        return jsonify({"error": "No files uploaded"}), 400

    files = request.files.getlist("files[]")
    if not files:
        return jsonify({"error": "No files received"}), 400

    all_rows = []
    all_columns_ordered = []
    seen_columns = set()
    results = []

    for file in files:
        if not file.filename.lower().endswith(".pdf"):
            results.append({"filename": file.filename, "status": "skipped", "reason": "Not a PDF"})
            continue

        try:
            pdf_bytes = file.read()
            # Extract text from all pages
            text = pdf_to_text(pdf_bytes)
            if not text.strip():
                results.append({"filename": file.filename, "status": "error", "reason": "No text extracted from PDF"})
                continue

            # Send to Claude for extraction
            rows = extract_comps_from_text(text)

            # Tag rows with source file and merge columns
            for row in rows:
                row["__source"] = file.filename
                for key in row.keys():
                    if key != "__source" and key not in seen_columns:
                        all_columns_ordered.append(key)
                        seen_columns.add(key)

            all_rows.extend(rows)
            results.append({
                "filename": file.filename,
                "status": "success",
                "rows_extracted": len(rows)
            })

        except Exception as e:
            results.append({
                "filename": file.filename,
                "status": "error",
                "reason": str(e)
            })

    return jsonify({
        "results": results,
        "total_rows": len(all_rows),
        "columns": all_columns_ordered,
        "rows": all_rows
    })


@app.route("/export", methods=["POST"])
def export():
    """
    Accepts JSON data (rows + columns), returns an Excel file download.
    Body: { "rows": [...], "columns": [...] }
    """
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    rows = data.get("rows", [])
    columns = data.get("columns", [])

    if not rows:
        return jsonify({"error": "No rows to export"}), 400

    try:
        wb = build_excel(rows, columns)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()

        filename = f"CRE_Comps_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return send_file(
            tmp.name,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
